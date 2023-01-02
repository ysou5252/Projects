import pandas as pd
import openpyxl
import xlsxwriter
import os
pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)


data_path_input = input('입력할 엑셀 파일 경로 + 파일명을 입력하세요: ')
print('\n')
data_path_output = input('저장할 엑셀 파일 경로 + 파일명을 입력하세요: ')
print('\n')
print('RT 회선 분류 진행중....')
print('\n')


# data_path = '/content/drive/MyDrive/Project/ETE 이원화 확인 Tool/E-T-E_tool_COT/TANGOT_EXCEL_DOWNLOAD_JOB202205030933.xlsx'


df3 = pd.read_excel(data_path_input, header=0,index_col=0)


df = df3.rename(columns=df3.iloc[0]).drop(df3.index[0])


with pd.ExcelWriter(data_path_output) as writer:
    df.to_excel(writer, sheet_name='WCDMA(NODEB)_1')


# data_path = '/content/drive/MyDrive/Project/ETE 이원화 확인 Tool/E-T-E_tool_COT/test1.xlsx'
df1 = pd.read_excel(data_path_output, header=0,index_col=0)


df_data = df1[['회선명','CUID','회선비고3','RT선번내용']] #원하는 Columns만 추출


CUID_yes = df_data[df_data['CUID'].notnull()] #CUID가 존재하는 index 추출 후 저장
CUID_no = df_data[df_data['CUID'].isnull()] #CUID가 존재하지 않는 index 추출 후 저장

notduplicateDFRow=CUID_yes.drop_duplicates(['CUID'], keep=False) #CUID 값 1개만 존재하는 index → <불량,조치필요#2>


two_over_duplicateDFRow = CUID_yes.drop(notduplicateDFRow.index) # 동일 CUID 2개 이상 존재 index


one_circuit_compare = two_over_duplicateDFRow.drop_duplicates(['회선비고3'], keep=False) # CUID 중복 2개이상 & 회선비고 값 1개 존재하는 index → <불량,조치필요3>


two_circuit_compare = two_over_duplicateDFRow.drop(one_circuit_compare.index) # CUID 중복 2개이상 & 회선비고 값 2개이상 → <양호>


total_data=pd.DataFrame({'기지국전체' : len(df1),
                         'RT 이원화 불량' : len(CUID_no.index)+len(notduplicateDFRow.index)+len(one_circuit_compare.index),
                         'CUID 미입력' : CUID_no['회선명'],
                         '수량#1' : len(CUID_no.index),
                         'CUID 단독입력' : notduplicateDFRow['회선명'],
                         '수량#2' : len(notduplicateDFRow.index),
                         '이원화불량' : one_circuit_compare['회선명'],
                         '수량#3' : len(one_circuit_compare.index)})


two_circuit_compare['장비명'] = two_circuit_compare['RT선번내용'].str.split('/').str[0] # <이원화 양호> 회선들의 RT 선번내용을 '/' 기준으로 text 분리 작업
two_circuit_compare['포트'] = two_circuit_compare['RT선번내용'].str.split('/').str[1] # <이원화 양호> 회선들의 RT 선번내용을 '/' 기준으로 text 분리 작업


column_null = two_circuit_compare[two_circuit_compare['RT선번내용'].isnull()] #<이원화 양호> 중 'RT선번내용'이 Null인 index

#len(column_null)

df_merge = pd.merge(two_circuit_compare, column_null, on='CUID', how='inner') #<이원화 양호>와 column_null의 inner join(CUID기준)

df_merge.rename(columns={'회선명_x':'회선명'},inplace=True)
df_merge.rename(columns={'회선비고3_x':'회선비고3'},inplace=True)
df_merge.rename(columns={'RT선번내용_x':'RT선번내용'},inplace=True)
df_merge.rename(columns={'장비명_x':'장비명'},inplace=True)
df_merge.rename(columns={'포트_x':'포트'},inplace=True)

df_merge = df_merge[df_merge.columns.drop(list(df_merge.filter(regex='_y')))]

df_merge = df_merge.drop_duplicates()


with pd.ExcelWriter(data_path_output) as writer:
    df.to_excel(writer, sheet_name='WCDMA(NODEB)_1')
    # df_COT_TID.to_excel(writer, sheet_name='COT_TID',index=False)
    total_data.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='RT 이원화 추출')
    CUID_no.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='CUID 미입력')
    notduplicateDFRow.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='CUID 단독입력')
    one_circuit_compare.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='이원화 불량')
    two_circuit_compare.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='RT 이원화 양호')
    df_merge.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='검증 필요')


from openpyxl import load_workbook
filename= data_path_output
wb = load_workbook(filename)
ws = wb.active
ws.insert_cols(4, 2)
ws.cell(row=1, column=4, value='COT 이원화 여부')
ws.cell(row=1, column=5, value='RT 이원화 여부')
wb.save(data_path_output)


# data_path = '/content/drive/MyDrive/Project/ETE 이원화 확인 Tool/E-T-E_tool_COT/test1.xlsx'
df_total = pd.read_excel(data_path_output, header=0, sheet_name = 'WCDMA(NODEB)_1')
df_total1 = df_total[['회선명', 'RT 이원화 여부']]


# data_path = '/content/drive/MyDrive/Project/ETE 이원화 확인 Tool/E-T-E_tool_COT/test1.xlsx'
CUID_no = pd.read_excel(data_path_output, header=0, sheet_name = 'CUID 미입력')
CUID_no1 = CUID_no[['회선명']]



df_merge1 = pd.merge(df_total1, CUID_no1)
df_merge1 = df_merge1.fillna('CUID 미입력')



# df_total['RT 이원화 여부'] = df_total['RT 이원화 여부'].fillna(df_merge['RT 이원화 여부'])
# df_total
df_total4 = pd.merge(df_total, df_merge1, how='left', on='회선명')
df_total4['RT 이원화 여부_x'] = df_total4['RT 이원화 여부_x'].fillna(df_total4['RT 이원화 여부_y'])



with pd.ExcelWriter(data_path_output) as writer:
    df_total4.to_excel(writer, sheet_name='WCDMA(NODEB)_1')
    # df_COT_TID.to_excel(writer, sheet_name='COT_TID',index=False)
    total_data.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='RT 이원화 추출')
    CUID_no.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='CUID 미입력')
    notduplicateDFRow.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='CUID 단독입력')
    one_circuit_compare.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='이원화 불량')
    two_circuit_compare.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='RT 이원화 양호')
    df_merge.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='검증 필요')


# data_path = '/content/drive/MyDrive/Project/ETE 이원화 확인 Tool/E-T-E_tool_COT/test1.xlsx'
df_total = pd.read_excel(data_path_output, header=0, sheet_name = 'WCDMA(NODEB)_1')
df_total.rename(columns = {'RT 이원화 여부_x':'RT 이원화 여부'},inplace=True)
df_total.drop(['RT 이원화 여부_y'], axis=1, inplace=True)
# df_total

with pd.ExcelWriter(data_path_output) as writer:
    df_total.to_excel(writer, sheet_name='WCDMA(NODEB)_1')
    # df_COT_TID.to_excel(writer, sheet_name='COT_TID',index=False)
    total_data.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='RT 이원화 추출')
    CUID_no.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='CUID 미입력')
    notduplicateDFRow.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='CUID 단독입력')
    one_circuit_compare.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='RT 이원화 불량')
    two_circuit_compare.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='RT 이원화 양호')
    df_merge.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='검증 필요')

    #조치필요#1 완료


# data_path = '/content/drive/MyDrive/Project/ETE 이원화 확인 Tool/E-T-E_tool_COT/test1.xlsx'
df_total = pd.read_excel(data_path_output, header=0, sheet_name = 'WCDMA(NODEB)_1')
df_total1 = df_total[['회선명', 'RT 이원화 여부']]


# data_path = '/content/drive/MyDrive/Project/ETE 이원화 확인 Tool/E-T-E_tool_COT/test1.xlsx'
notduplicateDFRow = pd.read_excel(data_path_output, header=0, sheet_name = 'CUID 단독입력')
notduplicateDFRow1 = notduplicateDFRow[['회선명']]


df_merge1= pd.merge(df_total1, notduplicateDFRow1)
df_merge1 = df_merge1.fillna('CUID 단독입력')


df_total4 = pd.merge(df_total, df_merge1, how='left', on='회선명')
df_total4['RT 이원화 여부_x'] = df_total4['RT 이원화 여부_x'].fillna(df_total4['RT 이원화 여부_y'])

with pd.ExcelWriter(data_path_output) as writer:
    df_total4.to_excel(writer, sheet_name='WCDMA(NODEB)_1')
    # df_COT_TID.to_excel(writer, sheet_name='COT_TID',index=False)
    total_data.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='RT 이원화 추출')
    CUID_no.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='CUID 미입력')
    notduplicateDFRow.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='CUID 단독입력')
    one_circuit_compare.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='이원화불량')
    two_circuit_compare.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='RT 이원화 양호')
    df_merge.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='검증 필요')


# data_path = '/content/drive/MyDrive/Project/ETE 이원화 확인 Tool/E-T-E_tool_COT/test1.xlsx'
df_total = pd.read_excel(data_path_output, header=0, sheet_name = 'WCDMA(NODEB)_1')
df_total.rename(columns = {'RT 이원화 여부_x':'RT 이원화 여부'},inplace=True)
df_total.drop(['RT 이원화 여부_y'], axis=1, inplace=True)
# df_total

with pd.ExcelWriter(data_path_output) as writer:
    df_total.to_excel(writer, sheet_name='WCDMA(NODEB)_1')
    # df_COT_TID.to_excel(writer, sheet_name='COT_TID',index=False)
    total_data.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='RT 이원화 추출')
    CUID_no.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='CUID 미입력')
    notduplicateDFRow.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='CUID 단독입력')
    one_circuit_compare.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='이원화불량')
    two_circuit_compare.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='RT 이원화 양호')
    df_merge.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='검증 필요')

    #조치필요#2 완료



# data_path = '/content/drive/MyDrive/Project/ETE 이원화 확인 Tool/E-T-E_tool_COT/test1.xlsx'
df_total = pd.read_excel(data_path_output, header=0, sheet_name = 'WCDMA(NODEB)_1')
df_total1 = df_total[['회선명', 'RT 이원화 여부']]


# data_path = '/content/drive/MyDrive/Project/ETE 이원화 확인 Tool/E-T-E_tool_COT/test1.xlsx'
one_circuit_compare = pd.read_excel(data_path_output, header=0, sheet_name = '이원화불량')
one_circuit_compare1 = one_circuit_compare[['회선명']]


df_merge1 = pd.merge(df_total1, one_circuit_compare1)
df_merge1 = df_merge1.fillna('이원화불량')


df_total4 = pd.merge(df_total, df_merge1, how='left', on='회선명')
df_total4['RT 이원화 여부_x'] = df_total4['RT 이원화 여부_x'].fillna(df_total4['RT 이원화 여부_y'])

with pd.ExcelWriter(data_path_output) as writer:
    df_total4.to_excel(writer, sheet_name='WCDMA(NODEB)_1')
    # df_COT_TID.to_excel(writer, sheet_name='COT_TID',index=False)
    total_data.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='RT 이원화 추출')
    CUID_no.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='CUID 미입력')
    notduplicateDFRow.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='CUID 단독입력')
    one_circuit_compare.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='이원화불량')
    two_circuit_compare.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='RT 이원화 양호')
    df_merge.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='검증 필요')


# data_path = '/content/drive/MyDrive/Project/ETE 이원화 확인 Tool/E-T-E_tool_COT/test1.xlsx'
df_total = pd.read_excel(data_path_output, header=0, sheet_name = 'WCDMA(NODEB)_1')
df_total.rename(columns = {'RT 이원화 여부_x':'RT 이원화 여부'},inplace=True)
df_total.drop(['RT 이원화 여부_y'], axis=1, inplace=True)
# df_total

with pd.ExcelWriter(data_path_output) as writer:
    df_total.to_excel(writer, sheet_name='WCDMA(NODEB)_1')
    # df_COT_TID.to_excel(writer, sheet_name='COT_TID',index=False)
    total_data.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='RT 이원화 추출')
    CUID_no.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='CUID 미입력')
    notduplicateDFRow.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='CUID 단독입력')
    one_circuit_compare.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='이원화불량')
    two_circuit_compare.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='RT 이원화 양호')
    df_merge.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='검증 필요')

    #조치필요#3 완료


# data_path = '/content/drive/MyDrive/Project/ETE 이원화 확인 Tool/E-T-E_tool_COT/test1.xlsx'
df_total = pd.read_excel(data_path_output, header=0, sheet_name = 'WCDMA(NODEB)_1')
df_total1 = df_total[['회선명', 'RT 이원화 여부']]


# data_path = '/content/drive/MyDrive/Project/ETE 이원화 확인 Tool/E-T-E_tool_COT/test1.xlsx'
two_circuit_compare = pd.read_excel(data_path_output, header=0, sheet_name = 'RT 이원화 양호')
two_circuit_compare1 = two_circuit_compare[['회선명']]


df_merge1 = pd.merge(df_total1, two_circuit_compare1)
df_merge1 = df_merge1.fillna('RT 이원화 양호')


df_total4 = pd.merge(df_total, df_merge1, how='left', on='회선명')
df_total4['RT 이원화 여부_x'] = df_total4['RT 이원화 여부_x'].fillna(df_total4['RT 이원화 여부_y'])
# df_total4
with pd.ExcelWriter(data_path_output) as writer:
    df_total4.to_excel(writer, sheet_name='WCDMA(NODEB)_1')
    # df_COT_TID.to_excel(writer, sheet_name='COT_TID',index=False)
    total_data.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='RT 이원화 추출')
    CUID_no.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='CUID 미입력')
    notduplicateDFRow.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='CUID 단독입력')
    one_circuit_compare.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='이원화불량')
    two_circuit_compare.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='RT 이원화 양호')
    df_merge.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='검증 필요')


# data_path = '/content/drive/MyDrive/Project/ETE 이원화 확인 Tool/E-T-E_tool_COT/test1.xlsx'
df_total = pd.read_excel(data_path_output, header=0, sheet_name = 'WCDMA(NODEB)_1')
df_total.rename(columns = {'RT 이원화 여부_x':'RT 이원화 여부'},inplace=True)
df_total.drop(['RT 이원화 여부_y'], axis=1, inplace=True)
# df_total

with pd.ExcelWriter(data_path_output) as writer:
    df_total.to_excel(writer, sheet_name='WCDMA(NODEB)_1')
    # df_COT_TID.to_excel(writer, sheet_name='COT_TID',index=False)
    total_data.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='RT 이원화 추출')
    CUID_no.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='CUID 미입력')
    notduplicateDFRow.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='CUID 단독입력')
    one_circuit_compare.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='이원화불량')
    two_circuit_compare.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='RT 이원화 양호')
    df_merge.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='검증 필요')

    #RT 이원화 양호 완료


# data_path = '/content/drive/MyDrive/Project/ETE 이원화 확인 Tool/E-T-E_tool_COT/test1.xlsx'
df_total = pd.read_excel(data_path_output, header=0, sheet_name = 'WCDMA(NODEB)_1')
df_total0 = df_total.drop(df_total.iloc[:,1:9], axis=1)

with pd.ExcelWriter(data_path_output) as writer:
    df_total0.to_excel(writer, sheet_name='WCDMA(NODEB)_1', index=False)
    # COT_TID_tolist1.to_excel(writer, sheet_name='COT_TID',index=False)
    total_data.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='RT 이원화 추출')
    CUID_no.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='CUID 미입력')
    notduplicateDFRow.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='CUID 단독입력')
    one_circuit_compare.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='이원화불량')
    two_circuit_compare.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='RT 이원화 양호')
    df_merge.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='검증 필요')


#병합시 필요없는 'unnamed' 컬럼 제거

# COT_TID = pd.read_excel(data_path_output, sheet_name='COT_TID', engine='openpyxl')

names= str(input("COT명을 입력하세요 : "))
print('\n')
print('COT 회선 분류 진행중....')
print('\n')

columns = ['COT_TID 입력값']
names1 = names.split(',')
result = [x.lstrip() for x in names1]
COT_TID_tolist1 = pd.DataFrame(result, columns = columns)


with pd.ExcelWriter(data_path_output) as writer:
    df_total0.to_excel(writer, sheet_name='WCDMA(NODEB)_1', index=False)
    COT_TID_tolist1.to_excel(writer, sheet_name='COT_TID',index=False)
    total_data.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='RT 이원화 추출')
    CUID_no.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='CUID 미입력')
    notduplicateDFRow.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='CUID 단독입력')
    one_circuit_compare.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='이원화불량')
    two_circuit_compare.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='RT 이원화 양호')
    df_merge.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='검증 필요')


COT_TID_tolist = COT_TID_tolist1['COT_TID 입력값'].tolist()
# COT_TID_tolist


two_circuit_compare['CUID'].tolist()


COT_CUID_ok = df_total0.loc[df_total0['CUID'].isin(two_circuit_compare['CUID'])]
# COT_CUID_ok


COT_CUID_ok1 = COT_CUID_ok.loc[:,'CUID':]
# COT_CUID_ok1


COT_TID_isin = COT_CUID_ok1.loc[:,"링#1":"EAST포트#12"][COT_CUID_ok1.loc[:,"링#1":"EAST포트#12"].isin(COT_TID_tolist)]
COT_CUID = COT_CUID_ok1['CUID']
COT_CUID1 = COT_CUID.astype(int)
COT_concat = pd.concat([COT_CUID1, COT_TID_isin], axis=1)


a= COT_concat.groupby('CUID').agg({lambda x: x.tolist()})
a1 = pd.DataFrame(a)
a1.columns = a1.columns.droplevel(1)


a2 = a1.applymap(lambda x: [*filter(pd.notna, x)])


a3 = a2.reset_index(level=0)


for col in a3.columns[1:]:
    for i in range(len(a3)):
        a_list = a3.loc[i,col]

        a_set = set(a_list)
        contains_duplicates = len(a_list) != len(a_set)
        a3.loc[i,col] = contains_duplicates


a4 = a3.set_index('CUID')


a4['COT 이원화 여부'] = a4.any(axis=1)
a5 = a4['COT 이원화 여부'].map({True: 'COT 미이원화', False: 'COT 이원화'})
a5 = pd.DataFrame(a5)


a6 = a5.reset_index(level=0)


a6['CUID']=a6['CUID'].astype(int)

big_df = df_total0.merge(a6, on='CUID', how='left')
big_df['COT 이원화 여부_x'] = big_df['COT 이원화 여부_x'].fillna(big_df['COT 이원화 여부_y'])
big_df.rename(columns = {'COT 이원화 여부_x':'COT 이원화 여부'},inplace=True)
big_df.drop(['COT 이원화 여부_y'], axis=1, inplace=True)

big_df['COT 이원화 여부'] = big_df['COT 이원화 여부'].fillna('CUID 등 확인필요')


with pd.ExcelWriter(data_path_output) as writer:
    big_df.to_excel(writer, sheet_name='WCDMA(NODEB)_1', index=False)
    COT_TID_tolist1.to_excel(writer, sheet_name='COT_TID',index=False)
    total_data.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='RT 이원화 추출')
    CUID_no.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='CUID 미입력')
    notduplicateDFRow.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='CUID 단독입력')
    one_circuit_compare.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='이원화불량')
    two_circuit_compare.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='RT 이원화 양호')
    df_merge.to_excel(writer, index=False, encoding = 'utf-8-sig', sheet_name='검증 필요')

#최종 COT,RT 완료 exe

writer.save()

os.system('pause')