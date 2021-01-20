from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "Joonyoung"

ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"]) #A1 셀의 정보를 출력
print(ws["A1"].value) #A1 셀의 값을 출력
print(ws["A10"].value) #값이 없을땐 none을 출력

print(ws.cell(row=1, column=1).value)


from random import *
index = 1
for x in range(1,11):
    for y in range(1,11):
        ws.cell(row=x, column=y, value=index)
        index += 1

wb.save("sample.xlsx")
